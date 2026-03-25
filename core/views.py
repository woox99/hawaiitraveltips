from urllib import response
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth import logout
from django.contrib.admin.views.decorators import staff_member_required
from django.urls import reverse



from core.models import *
from .util import *

from django.shortcuts import HttpResponse
from core.models import Booking, Island, Category
from openpyxl import load_workbook
from django.utils.text import slugify

from django.db.models import Q
import html



def index(request):
    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])

    # Get 4 popular bookings for the island
    bookings = Booking.objects.filter(is_pinned=True, is_public=True)[:4]
    
    context = {
        'island' : island,
        'islands' : Island.objects.all().order_by('modified'),
        'bookings': bookings,
    }
    return render(request, 'core/home.html', context)


def home(request):
    # Check if 'current_island' is in the session and render the island instead of the home page
    if 'current_island_slug' in request.session:
        return redirect('core:island', island_slug=request.session['current_island_slug'])

    # Get 4 popular bookings for the island
    bookings = Booking.objects.filter(is_pinned=True, is_public=True)[:4]
    
    context = {
        'islands' : Island.objects.all().order_by('modified'),
        'bookings': bookings,
    }
    return render(request, 'core/home.html', context)



def island_view(request, island_slug):

    # Set the current island in the session if coming from the home page
    island = get_object_or_404(Island, slug=island_slug)
    request.session['current_island_slug'] = island_slug

    # Get 4 popular bookings for the island
    bookings = Booking.objects.filter(island=island, is_public=True, is_pinned=True)[:4]

    # WordPress.com REST API endpoint - Limit = 3
    WP_API_URL = f"https://public-api.wordpress.com/wp/v2/sites/team92d3a5e49bc-kctlm.wordpress.com/posts?categories={island.wp_category_id}&per_page=3&_embed"
    wp_posts = get_wp_posts(WP_API_URL)

    # back_url = f'www.hawaiitraveltips.com/{quote(island.slug)}/?page={page_obj.number}'

    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
        'bookings': bookings,
        "wp_posts": wp_posts,
    }
    
    # Build the template path dynamically
    template_path = f'core/views/islands/{island.slug}.html'
    return render(request, template_path, context)



def bookings_view(request, island_slug):

    # Set the current island in the session if coming from the home page
    island = get_object_or_404(Island, slug=island_slug)
    request.session['current_island_slug'] = island_slug

    # Filter by category if category slug is provided in query params
    category_slug = request.GET.get('category', '')
    current_category = None
    if category_slug:
        for cat in Category.objects.all():
            if slugify(cat.name) == category_slug:
                current_category = cat
                break

    if current_category:
        bookings = Booking.objects.filter(island=island, tags=current_category).order_by("-is_pinned")
    elif request.GET.get('q'):
        query = request.GET.get('q')
        bookings = Booking.objects.filter(
            Q(title__icontains=query) |
            Q(city__icontains=query) |
            Q(details__icontains=query),
            island=island,
        ).order_by("-is_pinned")
    else:
        bookings = Booking.objects.filter(island=island).order_by("-is_pinned", "-is_popular", "-is_public")

    # Apply sorting based on query param
    sort_slug = request.GET.get('sort')

    if sort_slug:
        request.session['sort_slug'] = sort_slug
    else:
        sort_slug = request.session.get('sort_slug', '')

    if sort_slug == 'best-seller':
        bookings = bookings.order_by("-is_public", "-is_pinned", "-is_popular")
    elif sort_slug == 'rating':
        bookings = bookings.order_by("-company_rating", "-is_public")
    elif sort_slug == 'promo-code':
        bookings = bookings.order_by("-is_promo", "-promo_amount", 'is_public')
    elif sort_slug == 'price':
        bookings = bookings.order_by("-is_public","price", "-is_promo")
    elif sort_slug == 'title':
        bookings = bookings.order_by("-is_public", "title")

    if not request.user.is_authenticated:
        bookings = bookings.filter(is_public = True)

    # Pagination
    page_obj, page_range = paginate_bookings(bookings, request)

    # Get all categories that have visible bookings for this island
    categories = Category.objects.filter(bookings__island=island, bookings__is_public=True).distinct().order_by('name')
    
    # Get count of visible bookings for each category
    for cat in categories:
        # cat.visible_bookings_count = cat.bookings.filter(island=island).count
        cat.visible_bookings_count = cat.bookings.filter(island=island, is_public=True).count

    back_url = f'www.hawaiitraveltips.com/{island.slug}/tours-activities/?page={page_obj.number}&category={slugify(current_category)}'

        
    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
        'categories': categories,
        'current_category': current_category,
        'page_obj' : page_obj,
        'page_range': page_range,
        'bookings_count': bookings.count(),
        'total_bookings_count': Booking.objects.filter(island=island, is_public=True).count(),
        'sort_slug': sort_slug,
        'back_url' : back_url,
    }

    if request.GET.get('view'):
            request.session['view'] = request.GET.get('view')

    # Build the template path dynamically
    template_path = f'core/views/bookings/{island.slug}.html'
    return render(request, template_path, context)



def clear_bookings_filter(request, island_slug):
    if 'sort_slug' in request.session:
        del request.session['sort_slug']
    return redirect('core:bookings', island_slug)



def guide_list_view(request, island_slug):


    # set the current island in the session if coming from the home page
    island = get_object_or_404(Island, slug=island_slug)
    request.session['current_island_slug'] = island_slug

    # WordPress.com REST API endpoint
    WP_API_URL = f"https://public-api.wordpress.com/wp/v2/sites/team92d3a5e49bc-kctlm.wordpress.com/posts?categories={island.wp_category_id}&_embed"
    wp_posts = get_wp_posts(WP_API_URL)
    
    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
        "wp_posts": wp_posts,
    }

    # Build the template path dynamically
    template_path = f'core/views/guides/{island.slug}.html'
    return render(request, template_path, context)



def guide_detail_view(request, guide_slug):

    # WordPress.com REST API endpoint - filter by slug to get the specific post
    WP_API_URL = f"https://public-api.wordpress.com/wp/v2/sites/team92d3a5e49bc-kctlm.wordpress.com/posts?slug={guide_slug}&_embed"
    wp_posts = get_wp_posts(WP_API_URL)
    wp_post = wp_posts[0]

    ## If user is from site, get island_slug else if user is from search result get island from wp_post fetched
    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    else:
        wp_island_id = wp_post['_embedded']['wp:term'][0][0]['id']
        island = get_object_or_404(Island, wp_category_id=wp_island_id)
    # island_terms = wp_post['_embedded']['wp:term'][0]  # first list is categories
    # for term in island_terms:
    #     print(term['id'], term['slug'], term['name'])

    # Get all posts for the sidebar
    WP_API_URL = f"https://public-api.wordpress.com/wp/v2/sites/team92d3a5e49bc-kctlm.wordpress.com/posts?categories={island.wp_category_id}&_embed"
    wp_posts = get_wp_posts(WP_API_URL)

    title = html.unescape(wp_post['title']['rendered']).replace('\xa0', ' ').replace('&nbsp;', ' ')
    excerpt = html.unescape(wp_post['excerpt']['rendered']).replace('\xa0', ' ').replace('&nbsp;', ' ')
    
    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
        "wp_post": wp_post,
        "wp_posts": wp_posts,
        'title' : title,
        'excerpt' : excerpt,
    }

    return render(request, 'core/views/guide_detail.html', context)



def about_view(request):
    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    else:
        island = get_object_or_404(Island, slug='oahu')

    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
    }
    return render(request, 'core/views/about.html', context)



def contact_view(request):
    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    else:
        island = get_object_or_404(Island, slug='oahu')

    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
    }
    return render(request, 'core/views/contact.html', context)



def legal_view(request):
    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    else:
        island = get_object_or_404(Island, slug='oahu')

    context = {
        'island': island,
        'islands' : Island.objects.all().order_by('modified'),
    }
    return render(request, 'core/views/legal.html', context)


def logout_admin(request):

    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    logout(request)
    return redirect('core:bookings', island.slug)


@staff_member_required
def update_booking_view(request, booking_id):

    if 'current_island_slug' in request.session:
        island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    booking = get_object_or_404(Booking, pk=booking_id)


    context = {
        'booking' : booking,
        'island' : island,
        'islands' : Island.objects.all().order_by("modified"),
        'categories': Category.objects.all(),
        'current_category' : request.GET.get('category'),
        'current_page_num' : request.GET.get('page'),
    }

    return render(request, 'core/views/update.html', context)


@staff_member_required
def save_booking(request, booking_id):

    island = get_object_or_404(Island, slug=request.session['current_island_slug'])
        
    booking = update_booking_util(request, booking_id)


    context = {
        'booking' : booking,
        'island' : island,
        'islands' : Island.objects.all().order_by("modified"),
        'categories': Category.objects.all(),
        'current_category' : request.GET.get('category'),
        'current_page_num' : request.GET.get('page'),

    }
    return render(request, 'core/views/update.html', context)


@staff_member_required
def delete_booking(request, booking_id):
    # Get objects
    island = get_object_or_404(Island, slug=request.session['current_island_slug'])
    booking = get_object_or_404(Booking, pk=booking_id)

    # Delete booking
    booking.delete()

    # Get query params from request
    category = request.GET.get('category', '')
    page = request.GET.get('page', '')

    # Build URL
    url = reverse('core:bookings', kwargs={'island_slug': island.slug})
    if category or page:
        url += f'?category={category}&page={page}'

    return redirect(url)





# Importing from .xlsx file - debug
# Importing from .xlsx file - debug
# Importing from .xlsx file - debug
def to_bool(value):
    return str(value).strip().lower() in ["true", "1", "yes"]


def import_view(request):
    path = 'core/bookings2.xlsx'

    wb = load_workbook(path)
    sheet = wb.active

    # Get headers
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        try:
            # ISLAND
            island_name = (data.get("island") or "").strip()
            island, _ = Island.objects.get_or_create(name=island_name)

            # BOOKING (prevent duplicates using fh_id)
            booking, _ = Booking.objects.get_or_create(
                fh_id=data.get("fh_id"),
                defaults={
                    "title": data.get("title", ""),
                    "company_name": data.get("company_name", ""),
                    "company_rating": data.get("company_rating") or 0,
                    "company_reviews": data.get("company_reviews") or 0,
                    "city": data.get("city", ""),
                    "island": island,
                    "details": data.get("details", ""),
                    "duration": data.get("duration", ""),
                    "price": data.get("price") or 0,
                    "is_promo": to_bool(data.get("is_promo")),
                    "promo_amount": data.get("promo_amount") or 0,
                    "promo_code": data.get("promo_code", ""),
                    "is_public": to_bool(data.get("is_public")),
                    "is_popular": to_bool(data.get("is_popular")),
                    "is_pinned": to_bool(data.get("is_pinned")),
                    "referral_link": data.get("referral_link", ""),
                    "image_URL": data.get("image_URL", ""),
                }
            )

            # TAGS (handles: "Boat, Whale Watch")
            tags_string = data.get("tags", "")
            if tags_string:
                tag_names = [tag.strip() for tag in str(tags_string).split(",")]

                tag_objects = []
                for tag_name in tag_names:
                    if tag_name:
                        category, _ = Category.objects.get_or_create(name=tag_name)
                        tag_objects.append(category)

                booking.tags.set(tag_objects)

        except Exception as e:
            print(f"Skipping row due to error: {e}")

    return HttpResponse("Import complete ✅")

